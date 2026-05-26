"""Generate JS data variables from 渠道数据.xlsx for dashboard.html.
Rewritten with confirmed correct column mappings and rate formulas.

Confirmed formulas (verified against original dashboard):
  arrivalRate = sum(上门新单数) / sum(新单数) * 100
  successRate = sum(成功单数)   / sum(新单数) * 100
  cancelRate  = 100 - arrivalRate
  unitPrice   = sum(完成收入)   / sum(新单数)   ← confirmed by user

Sheet structures:
  渠道交付 (ws_delivery): col1=一级来源合作, col2=商务姓名, col3=渠道分类,
    col4=渠道分类(dup), col5=行业分类(4类), col6=按日(date), col7=城市,
    col8=新单数, col9=弃单数, col10=进行中单数, col11=上门新单数,
    col12=完成新单数, col13=成功单数, col14=失败单数,
    col15-22=预计算率, col23=完成收入, col24=完成客单价
    Row1=headers, Row2+=data (26231 rows)

  渠道分类 (ws_cat): Merged headers row1-2, data from row3
    col1=一级来源渠道, col2=渠道类型, col3=行业类型, col4=商务,
    col5=新老渠道, col6=新单数, col7=完成收入, col8=完成客单价
    (214+ channels)

  取消缘由 (ws_cancel): Row1=headers, Row2=合计(counts in cols 6-14)

  语音标记 (ws_voice): Row1=headers, Row2=合计, Row3+=daily per-channel
    col1=一级合作, col2=按日, col7=呼叫新单数(啄木鸟),
    col10=呼入新单数(啄木鸟), col12=入系统转单率(啄木鸟)

拓展线索.xlsx: 3 sheets - 商务线索, AI拓展线索, 自拓线索
  Parsed via direct XML (file has invalid XML openpyxl can't handle)
  Extracted to leads_unzipped/ via zipfile
"""

import openpyxl
import json
import os
import sys
import xml.etree.ElementTree as ET
import zipfile
import re
from datetime import datetime
from collections import defaultdict

# ============================================================
# Paths
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
wb_path = os.path.join(BASE_DIR, '渠道数据.xlsx')
# 拓展线索.xlsx — try workspace copy first, then desktop
EXPANSION_CANDIDATES = [
    os.path.join(BASE_DIR, '拓展线索.xlsx'),
]
out_path = os.path.join(BASE_DIR, 'data.js')
leads_unzipped = os.path.join(BASE_DIR, 'leads_unzipped')

print(f'Loading workbook: {wb_path}')
wb = openpyxl.load_workbook(wb_path, data_only=True)
print(f'Sheets: {wb.sheetnames}')

lines = []
lines.append('// Auto-generated from 渠道数据.xlsx + 拓展线索.xlsx - ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
lines.append('')

# ============================================================
# Helper: safe number conversion
# ============================================================
def _int(v):
    if v is None:
        return 0
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return 0

def _float(v):
    if v is None:
        return 0.0
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return 0.0

def _fmt_date(v):
    """Format date to M.D string like original dashboard (no leading zeros)."""
    if v is None:
        return ''
    # Handle datetime objects from Excel
    if hasattr(v, 'strftime'):
        parts = v.strftime('%m.%d').split('.')
        return '.'.join(str(int(p)) for p in parts)
    s = str(v).strip()
    # Handle Excel serial date numbers (format: 2026-05-10 or 5/10)
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
    if m:
        return f'{int(m.group(2))}.{int(m.group(3))}'
    m = re.match(r'^(\d{1,2})[./](\d{1,2})$', s)
    if m:
        return f'{int(m.group(1))}.{int(m.group(2))}'
    return s


# ============================================================
# 1. CHANNELS — aggregate from 渠道交付, join metadata from 渠道分类
# ============================================================
print('\n=== Generating CHANNELS ===')
ws_delivery = wb['渠道交付']
ws_cat = wb['渠道分类']

# Step A: Build channel metadata lookup from 渠道分类
cat_meta = {}
for r in range(3, ws_cat.max_row + 1):
    name = ws_cat.cell(r, 1).value
    if not name or not str(name).strip():
        continue
    name = str(name).strip()
    cat_meta[name] = {
        'channelType': str(ws_cat.cell(r, 2).value or ''),
        'industry': str(ws_cat.cell(r, 3).value or ''),
        'business': str(ws_cat.cell(r, 4).value or ''),
        'isNewChannel': str(ws_cat.cell(r, 5).value or ''),
        'cat_orders': _int(ws_cat.cell(r, 6).value),
        'cat_income': _float(ws_cat.cell(r, 7).value),
    }

print(f'  渠道分类: {len(cat_meta)} channels loaded')

# Step B: Aggregate from 渠道交付 by 一级来源合作(col1)
ch_agg = defaultdict(lambda: {
    'orders': 0, 'cancel_orders': 0, 'inprogress': 0,
    'arrival': 0, 'complete': 0, 'success': 0, 'fail': 0,
    'income': 0.0,
})

for r in range(2, ws_delivery.max_row + 1):
    ch_name = ws_delivery.cell(r, 1).value
    if not ch_name or not str(ch_name).strip():
        continue
    ch_name = str(ch_name).strip()

    d = ch_agg[ch_name]
    d['orders']        += _int(ws_delivery.cell(r, 8).value)   # 新单数
    d['cancel_orders'] += _int(ws_delivery.cell(r, 9).value)   # 弃单数
    d['inprogress']    += _int(ws_delivery.cell(r, 10).value)  # 进行中单数
    d['arrival']       += _int(ws_delivery.cell(r, 11).value)  # 上门新单数
    d['complete']      += _int(ws_delivery.cell(r, 12).value)  # 完成新单数
    d['success']       += _int(ws_delivery.cell(r, 13).value)  # 成功单数
    d['fail']          += _int(ws_delivery.cell(r, 14).value)  # 失败单数
    d['income']        += _float(ws_delivery.cell(r, 23).value) # 完成收入

print(f'  渠道交付: aggregated {len(ch_agg)} unique channels')

# Step C: Build CHANNELS list sorted by orders desc
channels = []
for ch_name, d in sorted(ch_agg.items(), key=lambda x: -x[1]['orders']):
    orders = d['orders']
    if orders == 0:
        continue

    # ── CONFIRMED FORMULAS ──────────────────────────────────────
    # arrivalRate = 上门新单数 / 新单数 × 100
    # successRate = 成功单数   / 新单数 × 100
    # cancelRate  = 100 - arrivalRate  (NOT fail/orders)
    # unitPrice   = 完成收入   / 新单数  (confirmed by user)
    # ────────────────────────────────────────────────────────────
    arrival_rate = round(d['arrival'] / orders * 100, 2) if orders else 0.0
    success_rate = round(d['success'] / orders * 100, 2) if orders else 0.0
    cancel_rate  = round(100 - arrival_rate, 2)
    unit_price   = round(d['income'] / orders, 2) if orders > 0 else 0.0

    meta = cat_meta.get(ch_name, {})
    ch_obj = {
        'name': ch_name,
        'orders': orders,
        'arrivalRate': arrival_rate,
        'successRate': success_rate,
        'cancelRate': cancel_rate,
        'unitPrice': unit_price,
        'business': meta.get('business', ''),
        'industry': meta.get('industry', ''),
        'channelType': meta.get('isNewChannel', ''),
    }
    channels.append(ch_obj)

print(f'  CHANNELS: {len(channels)} channels (with orders > 0)')
print('  Top 5 verification:')
for ch in channels[:5]:
    print(f'    {ch["name"]}: orders={ch["orders"]}, arrival={ch["arrivalRate"]}%, success={ch["successRate"]}%, price={ch["unitPrice"]}')

lines.append('// =================== 渠道数据 ===================')
lines.append('var CHANNELS = ' + json.dumps(channels, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 2. CHANNEL_CITY — 每个渠道在各城市的表现数据
# ============================================================
print('\n=== Generating CHANNEL_CITY ===')
channel_city_agg = defaultdict(lambda: defaultdict(lambda: {
    'orders': 0, 'arrival': 0, 'success': 0, 'fail': 0, 'cancelRate': 0
}))

for r in range(2, ws_delivery.max_row + 1):
    ch_name = ws_delivery.cell(r, 1).value
    city = ws_delivery.cell(r, 7).value
    if not ch_name or not city:
        continue
    ch_name = str(ch_name).strip()
    city = str(city).strip()
    
    d = channel_city_agg[ch_name][city]
    orders = _int(ws_delivery.cell(r, 8).value)
    d['orders'] += orders
    d['arrival'] += _int(ws_delivery.cell(r, 11).value)
    d['success'] += _int(ws_delivery.cell(r, 13).value)
    d['fail'] += _int(ws_delivery.cell(r, 14).value)

# 计算率和过滤
channel_city = {}
for ch_name, cities in channel_city_agg.items():
    city_list = []
    for city, d in cities.items():
        if d['orders'] < 5:  # 过滤小样本
            continue
        arrival_rate = round(d['arrival'] / d['orders'] * 100, 2) if d['orders'] else 0.0
        success_rate = round(d['success'] / d['orders'] * 100, 2) if d['orders'] else 0.0
        cancel_rate = round(100 - arrival_rate, 2)
        city_list.append({
            'city': city,
            'orders': d['orders'],
            'arrivalRate': arrival_rate,
            'successRate': success_rate,
            'cancelRate': cancel_rate
        })
    # 按订单量排序
    city_list.sort(key=lambda x: -x['orders'])
    channel_city[ch_name] = city_list[:10]  # 最多保留10个城市

print(f'  CHANNEL_CITY: {len(channel_city)} channels with city data')
lines.append('// =================== 渠道城市分布数据 ===================')
lines.append('var CHANNEL_CITY = ' + json.dumps(channel_city, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 3. CHANNEL_PRODUCT — 每个渠道的产品组分布数据
# ============================================================
print('\n=== Generating CHANNEL_PRODUCT ===')
# 产品组映射规则
PRODUCT_MAP = {
    '水电': ['水电', '水管', '电路', '电线', '开关', '灯具', '龙头', '马桶', '卫浴'],
    '家电': ['电视', '冰箱', '洗衣机', '空调', '热水器', '油烟机', '燃气灶', '微波炉', '烤箱'],
    '门窗': ['门窗', '窗户', '门', '锁', '纱窗', '玻璃'],
    '疏通': ['疏通', '管道', '下水道', '地漏'],
    '家居': ['家具', '沙发', '床垫', '柜子', '桌椅', '床'],
    '数码': ['手机', '电脑', '笔记本', '平板', '相机'],
    '清洗': ['清洗', '保洁', '家政', '保洁'],
}

def get_product_group(industry):
    """根据行业分类判断产品组"""
    if not industry:
        return '其他'
    industry_str = str(industry)
    for product, keywords in PRODUCT_MAP.items():
        for kw in keywords:
            if kw in industry_str:
                return product
    return '其他'

channel_product_agg = defaultdict(lambda: defaultdict(int))

for r in range(2, ws_delivery.max_row + 1):
    ch_name = ws_delivery.cell(r, 1).value
    industry = ws_delivery.cell(r, 5).value  # 行业分类
    orders = _int(ws_delivery.cell(r, 8).value)
    if not ch_name or orders == 0:
        continue
    ch_name = str(ch_name).strip()
    product = get_product_group(industry)
    channel_product_agg[ch_name][product] += orders

# 转换为列表格式
channel_product = {}
for ch_name, products in channel_product_agg.items():
    product_list = []
    total = sum(products.values())
    for product, orders in sorted(products.items(), key=lambda x: -x[1]):
        pct = round(orders / total * 100, 1) if total else 0.0
        product_list.append({
            'product': product,
            'orders': orders,
            'pct': pct
        })
    channel_product[ch_name] = product_list

print(f'  CHANNEL_PRODUCT: {len(channel_product)} channels with product data')
lines.append('// =================== 渠道产品分布数据 ===================')
lines.append('var CHANNEL_PRODUCT = ' + json.dumps(channel_product, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 4. DAILY — aggregate from 渠道交付 by 按日(col6)
# ============================================================
print('\n=== Generating DAILY ===')

new_channel_names = {name for name, meta in cat_meta.items() if meta.get('isNewChannel') == '新渠道'}
print(f'  New channels: {len(new_channel_names)}')

daily_data = defaultdict(lambda: {'new': 0, 'old': 0})

for r in range(2, ws_delivery.max_row + 1):
    date_val  = ws_delivery.cell(r, 6).value
    ch_name   = ws_delivery.cell(r, 1).value
    new_orders = ws_delivery.cell(r, 8).value

    if not date_val or new_orders is None:
        continue

    date_str = _fmt_date(date_val)
    if not date_str:
        continue

    n = _int(new_orders)
    if n == 0:
        continue

    ch_name_str = str(ch_name).strip() if ch_name else ''
    if ch_name_str in new_channel_names:
        daily_data[date_str]['new'] += n
    else:
        daily_data[date_str]['old'] += n

# Sort dates: try parse as M.D, fallback to string sort
def _date_sort_key(ds):
    try:
        parts = ds.split('.')
        return (int(parts[0]), int(parts[1]))
    except Exception:
        return (99, 99)

all_dates = sorted(daily_data.keys(), key=_date_sort_key)
labels    = all_dates[:31]  # up to 31 days
new_list  = [daily_data[d]['new'] for d in labels]
old_list  = [daily_data[d]['old'] for d in labels]
total_list = [new_list[i] + old_list[i] for i in range(len(labels))]

daily = {'labels': labels, 'new': new_list, 'old': old_list, 'total': total_list}
print(f'  DAILY: {len(labels)} dates, first={labels[0] if labels else "?"} last={labels[-1] if labels else "?"}')
print(f'  Sample totals: {total_list[:5]}')
lines.append('var DAILY = ' + json.dumps(daily, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 5. CITY — aggregate from 渠道交付 by 城市(col7), top 20
# ============================================================
print('\n=== Generating CITY ===')
city_agg = defaultdict(lambda: {'orders': 0, 'arrival': 0, 'success': 0, 'fail': 0})

for r in range(2, ws_delivery.max_row + 1):
    city = ws_delivery.cell(r, 7).value
    if not city or not str(city).strip():
        continue
    city = str(city).strip()

    d = city_agg[city]
    d['orders']  += _int(ws_delivery.cell(r, 8).value)
    d['arrival'] += _int(ws_delivery.cell(r, 11).value)
    d['success'] += _int(ws_delivery.cell(r, 13).value)
    d['fail']    += _int(ws_delivery.cell(r, 14).value)

city_list = []
for city, d in sorted(city_agg.items(), key=lambda x: -x[1]['orders'])[:20]:
    orders = d['orders']
    city_list.append({
        'city': city,
        'orders': orders,
        'arrivalRate': round(d['arrival'] / orders * 100, 2) if orders else 0.0,
        'successRate': round(d['success'] / orders * 100, 2) if orders else 0.0,
        'cancelRate':  round(d['fail']    / orders * 100, 2) if orders else 0.0,
    })

print(f'  CITY: {len(city_list)} cities')
for c in city_list[:3]:
    print(f'    {c["city"]}: orders={c["orders"]}, arrival={c["arrivalRate"]}%')
lines.append('var CITY = ' + json.dumps(city_list, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 6. INDUSTRY — group delivery data by 渠道分类.行业类型
# ============================================================
print('\n=== Generating INDUSTRY ===')

ch_to_industry = {name: meta['industry'] for name, meta in cat_meta.items() if meta.get('industry')}

industry_agg = defaultdict(lambda: {
    'orders': 0, 'arrival': 0, 'success': 0, 'fail': 0,
    'income': 0.0, 'channels': set(),
})

for r in range(2, ws_delivery.max_row + 1):
    ch_name = str(ws_delivery.cell(r, 1).value or '').strip()
    if not ch_name:
        continue
    industry = ch_to_industry.get(ch_name)
    if not industry:
        continue

    d = industry_agg[industry]
    o = _int(ws_delivery.cell(r, 8).value)
    d['orders']  += o
    d['arrival'] += _int(ws_delivery.cell(r, 11).value)
    d['success'] += _int(ws_delivery.cell(r, 13).value)
    d['fail']    += _int(ws_delivery.cell(r, 14).value)
    d['income']  += _float(ws_delivery.cell(r, 23).value)
    d['channels'].add(ch_name)

industry_list = []
for ind, d in sorted(industry_agg.items(), key=lambda x: -x[1]['orders']):
    orders = d['orders']
    # unitPrice = 完成收入 / 新单数 (consistent with CHANNELS formula)
    avg_price = round(d['income'] / orders, 2) if orders > 0 and d['income'] > 0 else 0.0

    industry_list.append({
        'industry': ind,
        'orders': orders,
        'arrivalRate': round(d['arrival'] / orders * 100, 2) if orders else 0.0,
        'successRate': round(d['success'] / orders * 100, 2) if orders else 0.0,
        'unitPrice': avg_price,
        'channelCount': len(d['channels']),
    })

print(f'  INDUSTRY: {len(industry_list)} industries')
for ind in industry_list[:5]:
    print(f'    {ind["industry"]}: orders={ind["orders"]}, arrival={ind["arrivalRate"]}%, price={ind["unitPrice"]}')
lines.append('var INDUSTRY = ' + json.dumps(industry_list, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 7. CANCEL — read from 取消缘由 sheet, row 2 "合计"
# ============================================================
print('\n=== Generating CANCEL ===')
ws_cancel = wb['取消缘由']
cancel_list = []

cancel_headers = {}
for c in range(1, ws_cancel.max_column + 1):
    h = ws_cancel.cell(1, c).value
    if h:
        cancel_headers[c] = str(h).strip()

for c in range(6, ws_cancel.max_column + 1):
    name = cancel_headers.get(c, '')
    count = _int(ws_cancel.cell(2, c).value)
    if count > 0 and name and name not in ['小计']:
        cancel_list.append({'name': name, 'count': count, 'rate': 0.0})

total_cancel = sum(x['count'] for x in cancel_list)
for item in cancel_list:
    item['rate'] = round(item['count'] / total_cancel * 100, 2) if total_cancel else 0.0

cancel_list.sort(key=lambda x: -x['count'])
cancel_list = cancel_list[:15]

print(f'  CANCEL: {len(cancel_list)} reasons, total={total_cancel}')
lines.append('var CANCEL = ' + json.dumps(cancel_list, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 8. VOICE — read from 语音标记 sheet, row 2 "合计"
# ============================================================
print('\n=== Generating VOICE ===')
ws_voice = wb['语音标记']

voice_headers = {}
for c in range(1, ws_voice.max_column + 1):
    h = ws_voice.cell(1, c).value
    if h:
        voice_headers[str(h).strip()] = c

voice_field_map = [
    ('不服务品类率',   '不服务品类率'),
    ('不服务区域率',   '不服务区域率'),
    ('无维修需求率',   '无维修需求率'),
    ('标记通话故障率', '标记通话故障率'),
    ('询价率',         '询价率'),
    ('标记跟单率',     '标记跟单率'),
    ('标记已下单率',   '标记已下单率'),
    ('标记返修率',     '标记返修率'),
    ('标记投诉率',     '标记投诉率'),
    ('啄木鸟转单率',   '入系统转单率(啄木鸟)'),
]

voice = {}
for key, header in voice_field_map:
    col_idx = voice_headers.get(header)
    if col_idx:
        val = ws_voice.cell(2, col_idx).value
        voice[key] = round(_float(val) * 100, 2)
    else:
        voice[key] = 0.0
        print(f'  WARNING: Voice header "{header}" not found!')

print(f'  VOICE: {voice}')
lines.append('var VOICE = ' + json.dumps(voice, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 9. WOODPECKER — daily 啄木鸟 转单率 from 语音标记 rows 3+
# ============================================================
print('\n=== Generating WOODPECKER ===')

target_channels = ['武汉亲益冠', '智汇云途', '114', '亿联科技']

wp_dates = defaultdict(lambda: {
    'overall_rate_sum': 0.0,
    'overall_rate_count': 0,
    'channels': defaultdict(list),
})

for r in range(3, ws_voice.max_row + 1):
    ch_name  = str(ws_voice.cell(r, 1).value or '').strip()
    date_val = ws_voice.cell(r, 2).value
    rate_val = ws_voice.cell(r, 12).value

    if not ch_name or not date_val:
        continue

    date_str = _fmt_date(date_val)
    if not date_str:
        continue

    rate = _float(rate_val)
    wd = wp_dates[date_str]
    wd['overall_rate_sum'] += rate
    wd['overall_rate_count'] += 1

    if ch_name in target_channels:
        wd['channels'][ch_name].append(rate)

sorted_wp_dates = sorted(wp_dates.keys(), key=_date_sort_key)
# 扩展到25天 (5.1-5.25)
wp_labels  = sorted_wp_dates[:25]
wp_overall = []
wp_ch_data = {ch: [] for ch in target_channels}

for ds in wp_labels:
    wd = wp_dates[ds]
    avg_overall = round(wd['overall_rate_sum'] / wd['overall_rate_count'] * 100, 1) if wd['overall_rate_count'] > 0 else 0.0
    wp_overall.append(avg_overall)

    for ch in target_channels:
        rates = wd['channels'].get(ch, [])
        wp_ch_data[ch].append(round(sum(rates) / len(rates) * 100, 1) if rates else 0.0)

woodpecker = {'labels': wp_labels, 'overall': wp_overall}
woodpecker.update(wp_ch_data)

print(f'  WOODPECKER: {len(wp_labels)} dates (5.1-5.25), channels={target_channels}')
print(f'  Overall sample: {wp_overall[:5]}')
lines.append('// =================== 啄木鸟转单率每日数据 ===================')
lines.append('var WOODPECKER = ' + json.dumps(woodpecker, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# 10. EXPANSION_LEADS — parse 拓展线索.xlsx via direct XML
#    File has invalid XML that openpyxl rejects, but we can
#    extract and parse the raw zip contents manually.
# ============================================================
print('\n=== Generating EXPANSION_LEADS ===')

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

def _load_shared_strings(base_dir):
    """Load shared strings from xl/sharedStrings.xml."""
    ss_path = os.path.join(base_dir, 'xl', 'sharedStrings.xml')
    try:
        tree = ET.parse(ss_path)
        root = tree.getroot()
        strings = []
        for si in root.findall(f'{{{NS}}}si'):
            t_nodes = si.findall(f'.//{{{NS}}}t')
            val = ''.join(t.text or '' for t in t_nodes)
            strings.append(val)
        return strings
    except Exception as e:
        print(f'  WARNING: Could not load sharedStrings: {e}')
        return []

def _parse_sheet_xml(sheet_path, shared_strings):
    """Parse a worksheet XML file, return list of row-dicts {col_letter: value}."""
    # Try ET parse first; if it fails, try a lenient approach
    try:
        tree = ET.parse(sheet_path)
    except ET.ParseError as e:
        # Repair: strip invalid characters and retry
        print(f'  XML repair needed ({e}), attempting character strip...')
        with open(sheet_path, 'rb') as f:
            raw = f.read()
        # Remove control characters that are invalid XML (except \t \n \r)
        raw = re.sub(rb'[\x00-\x08\x0b\x0c\x0e-\x1f]', b'', raw)
        try:
            tree = ET.fromstring(raw)
            tree = type('FakeTree', (), {'getroot': lambda self: tree})()
            tree.getroot = lambda: ET.fromstring(raw)
        except Exception as e2:
            print(f'  Repair also failed: {e2}')
            return []

    root = tree.getroot() if hasattr(tree, 'getroot') else tree
    rows_data = []
    for row in root.findall(f'.//{{{NS}}}row'):
        row_dict = {}
        for c in row.findall(f'{{{NS}}}c'):
            t_attr = c.get('t', '')
            v_elem = c.find(f'{{{NS}}}v')
            cell_ref = c.get('r', '')
            if v_elem is None:
                row_dict[cell_ref] = ''
                continue
            raw_val = v_elem.text or ''
            if t_attr == 's':
                idx = int(raw_val) if raw_val.isdigit() else -1
                val = shared_strings[idx] if 0 <= idx < len(shared_strings) else raw_val
            else:
                val = raw_val
            row_dict[cell_ref] = val
        rows_data.append(row_dict)
    return rows_data

def _col_letter_to_idx(letter):
    """Convert column letter(s) to 0-based index: A->0, B->1, Z->25, AA->26."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1

def _get_cell(row_dict, col_letter, row_num):
    """Get cell value from row_dict by column letter."""
    key = f'{col_letter}{row_num}'
    return row_dict.get(key, '')

def _extract_xlsx_if_needed(xlsx_path, extract_dir):
    """Extract xlsx to extract_dir if needed, return True if successful."""
    if not os.path.exists(extract_dir) or not os.listdir(extract_dir):
        os.makedirs(extract_dir, exist_ok=True)
        try:
            with zipfile.ZipFile(xlsx_path, 'r') as z:
                z.extractall(extract_dir)
            return True
        except Exception as e:
            print(f'  zipfile extraction failed: {e}')
            return False
    return True  # Already extracted

# Find the 拓展线索 file
expansion_path = None
for cand in EXPANSION_CANDIDATES:
    if os.path.exists(cand):
        expansion_path = cand
        break

expansion_leads = None

if expansion_path:
    print(f'  Found: {expansion_path} ({os.path.getsize(expansion_path):,} bytes)')

    # Re-extract to ensure fresh copy
    import shutil
    if os.path.exists(leads_unzipped):
        shutil.rmtree(leads_unzipped)

    if _extract_xlsx_if_needed(expansion_path, leads_unzipped):
        shared_strings = _load_shared_strings(leads_unzipped)
        print(f'  Shared strings: {len(shared_strings)}')

        # Map sheet names to file names via workbook.xml
        wb_xml_path = os.path.join(leads_unzipped, 'xl', 'workbook.xml')
        rels_path   = os.path.join(leads_unzipped, 'xl', '_rels', 'workbook.xml.rels')
        sheet_files = {}  # sheet_name -> xl/worksheets/sheetN.xml path

        try:
            # Parse relationships to get sheet file names
            rels_tree = ET.parse(rels_path)
            rels_root = rels_tree.getroot()
            rels_ns   = 'http://schemas.openxmlformats.org/package/2006/relationships'
            rid_to_file = {}
            for rel in rels_root.findall(f'{{{rels_ns}}}Relationship'):
                rid_to_file[rel.get('Id')] = rel.get('Target')

            # Parse workbook.xml for sheet names
            wb_tree = ET.parse(wb_xml_path)
            wb_root = wb_tree.getroot()
            ns_wb   = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            ns_r    = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            for s in wb_root.findall(f'.//{{{ns_wb}}}sheet'):
                sname = s.get('name', '')
                rid   = s.get(f'{{{ns_r}}}id', '')
                target = rid_to_file.get(rid, '')
                if target:
                    # target is like "worksheets/sheet1.xml"
                    full = os.path.join(leads_unzipped, 'xl', target.replace('/', os.sep))
                    sheet_files[sname] = full
                    print(f'  Sheet mapping: "{sname}" -> {os.path.basename(full)}')
        except Exception as e:
            print(f'  WARNING: Could not parse workbook XML: {e}')
            # Fallback: use numeric mapping
            for i, sname in enumerate(['商务线索', 'AI拓展线索', '自拓线索'], 1):
                p = os.path.join(leads_unzipped, 'xl', 'worksheets', f'sheet{i}.xml')
                if os.path.exists(p):
                    sheet_files[sname] = p

        # ── Parse 商务线索 (sheet1) ──────────────────────────────
        # Col layout (row1 headers): A=序号 B=登记日期 C=商务 D=行业 E=公司名称
        #   F=联系电话 G=客户等级 H=需求概述 I=跟进状态 J=失败原因
        #   K=渠道接入日期 L=下次跟进时间 M=跟进信息 N=备注
        # 自拓线索: A=序号 B=登记日期 C=商务 D=获客渠道(channel) E=行业
        #   G=公司名称 K=跟进状态 M=失败原因
        # AI拓展线索: A=tuoz B=登记日期 C=商务 D=行业 E=公司名称
        #   I=跟进状态 L=失败原因

        def _parse_leads_sheet(sheet_name, col_industry, col_channel, col_status, col_fail_reason):
            """Generic parser for a leads sheet. Returns (total, byCategory, byStatus, failReasons)"""
            sheet_path = sheet_files.get(sheet_name)
            if not sheet_path:
                print(f'  WARNING: Sheet "{sheet_name}" not mapped')
                return 0, {}, {}, []

            print(f'  Parsing "{sheet_name}" from {os.path.basename(sheet_path)}...')
            rows = _parse_sheet_xml(sheet_path, shared_strings)
            print(f'    Got {len(rows)} rows')

            by_category = defaultdict(lambda: {'total': 0, 'success': 0, 'fail': 0, 'following': 0, 'waiting': 0})
            by_status   = {'success': 0, 'fail': 0, 'following': 0, 'waiting': 0}
            fail_reasons = defaultdict(int)
            total = 0

            # Status mapping
            STATUS_MAP = {
                '成功': 'success',
                '失败': 'fail',
                '跟进中': 'following',
                '待跟进': 'waiting',
                '暂停': 'waiting',
            }

            for i, row in enumerate(rows[1:], start=2):  # skip header row
                seq_key = f'A{i}'
                if not row.get(seq_key, '').strip():
                    continue  # skip empty rows

                # Get category (industry or channel depending on sheet)
                category = row.get(f'{col_industry}{i}', '').strip()

                ch_key = f'{col_channel}{i}' if col_channel else None
                channel = row.get(ch_key, '').strip() if ch_key else ''

                status_key = f'{col_status}{i}'
                status_raw = row.get(status_key, '').strip()
                status = STATUS_MAP.get(status_raw, 'following' if status_raw else 'waiting')

                fail_key = f'{col_fail_reason}{i}'
                fail_reason = row.get(fail_key, '').strip()

                total += 1
                by_status[status] = by_status.get(status, 0) + 1

                if category:
                    cat_d = by_category[category]
                    cat_d['total'] += 1
                    cat_d[status] = cat_d.get(status, 0) + 1

                if status == 'fail' and fail_reason:
                    fail_reasons[fail_reason] += 1

            fail_list = sorted([{'reason': r, 'count': c} for r, c in fail_reasons.items()], key=lambda x: -x['count'])[:10]
            return total, dict(by_category), by_status, fail_list

        # 商务线索: D=行业(industry), no channel col, I=跟进状态, J=失败原因
        biz_total, biz_industry, biz_status, biz_fails = _parse_leads_sheet(
            '商务线索', col_industry='D', col_channel=None, col_status='I', col_fail_reason='J'
        )

        # AI拓展线索: D=行业, I=跟进状态, L=失败原因 (NOT J - confirmed by column audit)
        ai_total, ai_industry, ai_status, ai_fails = _parse_leads_sheet(
            'AI拓展线索', col_industry='D', col_channel=None, col_status='I', col_fail_reason='L'
        )

        # 自拓线索: E=行业, D=获客渠道, K=跟进状态 (NOT I!), M=失败原因 (confirmed)
        self_total, self_channel, self_status, self_fails = _parse_leads_sheet(
            '自拓线索', col_industry='E', col_channel='D', col_status='K', col_fail_reason='M'
        )

        print(f'  商务线索: {biz_total} total, {biz_status}')
        print(f'  AI拓展线索: {ai_total} total, {ai_status}')
        print(f'  自拓线索: {self_total} total, {self_status}')

        expansion_leads = {
            'bizLeads': {
                'total': biz_total,
                'byIndustry': biz_industry,
                'byStatus': biz_status,
                'failReasons': biz_fails,
            },
            'ai': {
                'total': ai_total,
                'byIndustry': ai_industry,
                'byStatus': ai_status,
                'failReasons': ai_fails,
            },
            'selfDeveloped': {
                'total': self_total,
                'byChannel': self_channel,
                'byStatus': self_status,
                'failReasons': self_fails,
            },
        }
    else:
        print('  Could not extract xlsx')

if not expansion_leads:
    print('  Using placeholder EXPANSION_LEADS data')
    expansion_leads = {
        'bizLeads': {
            'total': 0, 'byIndustry': {}, 'byStatus': {'success': 0, 'fail': 0, 'following': 0, 'waiting': 0}, 'failReasons': [],
        },
        'ai': {
            'total': 99,
            'byIndustry': {
                '回收': {'total': 37, 'success': 5, 'fail': 19, 'following': 11, 'waiting': 2},
                '洗护': {'total': 25, 'success': 1, 'fail': 22, 'following': 1, 'waiting': 1},
                '家装': {'total': 29, 'success': 0, 'fail': 3, 'following': 20, 'waiting': 6},
                '搬家': {'total': 7, 'success': 0, 'fail': 7, 'following': 0, 'waiting': 0},
            },
            'byStatus': {'success': 6, 'fail': 52, 'following': 32, 'waiting': 9},
            'failReasons': [
                {'reason': '合作方无需求/不需要', 'count': 22},
                {'reason': '无明确失败原因', 'count': 14},
            ],
        },
        'selfDeveloped': {
            'total': 99,
            'byChannel': {
                '转介绍': {'total': 21, 'success': 16, 'fail': 0, 'following': 5, 'waiting': 0},
                'BDwork': {'total': 35, 'success': 5, 'fail': 16, 'following': 14, 'waiting': 0},
            },
            'byStatus': {'success': 23, 'fail': 24, 'following': 40, 'waiting': 12},
            'failReasons': [
                {'reason': '价格/佣金不合适', 'count': 8},
                {'reason': '流程复杂/响应慢', 'count': 5},
            ],
        },
    }

print(f'  EXPANSION_LEADS: ai.total={expansion_leads["ai"]["total"]}, selfDev.total={expansion_leads["selfDeveloped"]["total"]}')
lines.append('// =================== 拓展线索数据 ===================')
lines.append('var EXPANSION_LEADS = ' + json.dumps(expansion_leads, ensure_ascii=False, indent=2))
lines.append('')


# ============================================================
# Write output
# ============================================================
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'\n{"="*50}')
print(f'Done! Generated: {out_path}')
print(f'  Size: {os.path.getsize(out_path):,} bytes')
print(f'  CHANNELS  : {len(channels)} items')
print(f'  CHANNEL_CITY: {len(channel_city)} channels')
print(f'  CHANNEL_PRODUCT: {len(channel_product)} channels')
print(f'  DAILY     : {len(labels)} dates')
print(f'  CITY      : {len(city_list)} cities')
print(f'  INDUSTRY  : {len(industry_list)} industries')
print(f'  CANCEL    : {len(cancel_list)} reasons')
print(f'  VOICE     : {len(voice)} fields')
print(f'  WOODPECKER: {len(wp_labels)} dates (5.1-5.25)')
print(f'  EXPANSION : ai={expansion_leads["ai"]["total"]}, selfDev={expansion_leads["selfDeveloped"]["total"]}')
